
import os
from datetime import datetime
from pathlib import Path
import json
import pandas as pd
from openpyxl import load_workbook, Workbook
from sqlalchemy import create_engine, Column, Integer, String, JSON
from sqlalchemy.orm import sessionmaker, declarative_base

# ==================== 配置数据库连接 ====================
# 根据 Flask 配置中的 SQLALCHEMY_BINDS 设置实际路径
DATABASES = {
    'navigation': 'sqlite:///instance/navigation.db',   # Navigation 模型对应的数据库
    'service': 'sqlite:///instance/service.db'          # Service 模型对应的数据库
}

# ==================== 定义模型（与 Flask 模型完全一致）====================
Base = declarative_base()

class Navigation(Base):
    __tablename__ = 'navigation'
    id = Column(Integer, primary_key=True)
    name = Column(String(80), nullable=False, unique=True)   # name 唯一
    description = Column(String(200), nullable=False)
    icon = Column(String(200), nullable=False)
    url = Column(String(200), nullable=False, unique=True)   # url 唯一
    group = Column(String(80), nullable=False)
    tags = Column(JSON, nullable=False)

class Service(Base):
    __tablename__ = 'service'
    id = Column(Integer, primary_key=True)
    name = Column(String(80), nullable=False)                # name 不唯一
    description = Column(String(200), nullable=False)
    icon = Column(String(200), nullable=False)
    url = Column(String(200), nullable=False, unique=True)   # url 唯一
    group = Column(String(80), nullable=False)
    tags = Column(JSON, nullable=False)

# 工作表名到模型类的映射
MODEL_MAP = {
    'Navigation': Navigation,
    'Service': Service
}

# ==================== 工具函数 ====================
def parse_tags(tags_value):
    """将 tags 字段转换为 Python 列表"""
    if pd.isna(tags_value):
        return []
    if isinstance(tags_value, str):
        # 尝试解析为 JSON
        try:
            return json.loads(tags_value)
        except json.JSONDecodeError:
            # 按逗号分隔处理
            return [tag.strip() for tag in tags_value.split(',') if tag.strip()]
    # 如果已经是列表，直接返回
    return tags_value

def get_database_engine(bind_key):
    """根据绑定键获取数据库引擎，并创建表（如果不存在）"""
    db_uri = DATABASES.get(bind_key)
    if not db_uri:
        raise ValueError(f"未找到绑定键 '{bind_key}' 对应的数据库 URI")
    engine = create_engine(db_uri)
    # 创建对应模型的表
    if bind_key == 'navigation':
        Navigation.metadata.create_all(engine)
    elif bind_key == 'service':
        Service.metadata.create_all(engine)
    return engine

def insert_data(session, model_class, data_rows):
    """
    插入数据到指定模型，根据模型的唯一约束检查重复。
    - Navigation: name 和 url 都唯一，需检查两者是否已存在
    - Service: 仅 url 唯一，只需检查 url
    """
    for row in data_rows:
        # 根据模型类型构建唯一性检查条件
        if model_class == Navigation:
            existing = session.query(model_class).filter(
                (model_class.name == row['name']) | (model_class.url == row['url'])
            ).first()
        elif model_class == Service:
            existing = session.query(model_class).filter(
                model_class.url == row['url']
            ).first()
        else:
            existing = None  # 不会发生

        if existing:
            print(f"跳过重复记录: name={row['name']}, url={row['url']} (表 {model_class.__tablename__})")
            continue

        # 创建实例并添加
        instance = model_class(
            name=row['name'],
            description=row['description'],
            icon=row['icon'],
            url=row['url'],
            group=row['group'],
            tags=row['tags']
        )
        session.add(instance)

    session.commit()
    print(f"成功导入 {len(data_rows)} 条记录到 {model_class.__tablename__} 表")


def insert_columns(sheetOriginalName, sheetNewName, skip_header=True):
    # 写入表头
    headers = ['id', 'name', 'description', 'icon', 'url', 'group', 'tags']
    for col_idx, value in enumerate(headers, start=1):
        sheetNewName.cell(row=1, column=col_idx, value=value)

    # 确定数据范围（从第几行开始）
    start_row = 2 if skip_header else 1
    max_row = sheetOriginalName.max_row

    # 逐行读取源数据（从第 2 行开始），写入目标（从第 2 行开始）
    for row_idx in range(start_row, max_row + 1):
        row_id = start_row - 1
        # 将自增 ID 写入目标 A 列
        sheetNewName.cell(row=start_row, column=1, value=row_id)
        # 读取源数据 A~E 列（列号 1~5），写入目标 B~F 列（列号 2~6）
        for col_idx in range(1, 6):  # 1:A, 2:B, ..., 5:E
            src_val = sheetOriginalName.cell(row=row_idx, column=col_idx).value
            sheetNewName.cell(row=row_idx, column=col_idx+1, value=src_val)
        # 读取源数据 F~Q 列（列号 6~17），处理后写入目标 G 列（列号 7）
        merged_values = []
        for col_idx in range(6, 18):  # openpyxl 列号从 1 开始
            cell_value = str(sheetOriginalName.cell(row=row_idx, column=col_idx).value)
            if (col_idx == 6) and (cell_value == "•"):       # 行政部门
                merged_values.append("ADMINISTRATION")
            elif (col_idx == 7) and (cell_value == "•"):     # 动画部门
                merged_values.append("ANIMATION")
            elif (col_idx == 8) and (cell_value == "•"):     # 董事部门
                merged_values.append("BOARD")
            elif (col_idx == 9) and (cell_value == "•"):     # 商务部门
                merged_values.append("BUSINESS")
            elif (col_idx == 10) and (cell_value == "•"):    # 设计部门
                merged_values.append("DESIGN")
            elif (col_idx == 11) and (cell_value == "•"):    # 开发部门
                merged_values.append("DEVELOPMENT")
            elif (col_idx == 12) and (cell_value == "•"):    # 特效部门
                merged_values.append("EFFECTS")
            elif (col_idx == 13) and (cell_value == "•"):    # 财务部门
                merged_values.append("FINANCE")
            elif (col_idx == 14) and (cell_value == "•"):    # 建模部门
                merged_values.append("MODELING")
            elif (col_idx == 15) and (cell_value == "•"):    # 运维部门
                merged_values.append("OPERATION")
            elif (col_idx == 16) and (cell_value == "•"):    # 摄影部门
                merged_values.append("PHOTOGRAPHY")
            elif (col_idx == 17) and (cell_value == "•"):    # 视频部门
                merged_values.append("VIDEO")

        # 合并所有值（可根据需要修改分隔符）
        merged_str = ",".join(merged_values)
        # 将合并后的字符串写入写入目标 G 列（列号 7）
        sheetNewName.cell(row=start_row, column=7, value=merged_str)

        start_row += 1

def merge_columns_to_new_file(input_file, output_file):
    """
    读取 input_file 中的 G~R 列数据，逐行合并后写入 output_file 的 G 列

    :param input_file:  源表格 A 的文件路径
    :param output_file: 目标表格 B 的文件路径
    :param skip_header: 是否跳过第一行（表头），默认 True
    """
    # 加载源工作簿
    originalWorkbook = load_workbook(input_file)
    sheetOriginalNavigation = originalWorkbook['Navigation']
    sheetOriginalService = originalWorkbook['Service']

    # 创建新工作簿用于写入结果
    newWorkbook = Workbook()
    # 获取默认的工作表
    sheetNewNavigation = newWorkbook.active
    # 设置工作表名称
    sheetNewNavigation.title = "Navigation"
    # 新建一个工作表，命名为 "Service"
    sheetNewService = newWorkbook.create_sheet("Service")

    insert_columns(sheetOriginalNavigation, sheetNewNavigation, )
    insert_columns(sheetOriginalService, sheetNewService, )

    newWorkbook.save(output_file)       # 保存结果工作簿
    originalWorkbook.close()            # 关闭源文件释放资源


# ==================== 主处理函数 ====================
def excel_to_sqlite(excel_file):
    """
    读取 Excel 文件，根据工作表名分别写入对应的数据库。
    """
    output_file = r"./instance/data-sqlite.xlsx"
    if os.path.exists(output_file):
        os.remove(output_file)

    merge_columns_to_new_file(excel_file, output_file)

    # 写入数据库
    xls = pd.ExcelFile(output_file)
    for sheet_name in xls.sheet_names:
        if sheet_name not in MODEL_MAP:
            print(f"跳过未知工作表: {sheet_name}")
            continue

        model_class = MODEL_MAP[sheet_name]
        bind_key = model_class.__tablename__  # 绑定键与表名相同
        print(f"处理工作表: {sheet_name} -> 数据库: {bind_key}")

        # 读取工作表数据
        df = pd.read_excel(xls, sheet_name=sheet_name)

        # 检查必需的列
        required_columns = ["id", "name", "description", "icon", "url", "group", "tags"]
        missing = [col for col in required_columns if col not in df.columns]
        if missing:
            raise ValueError(f"工作表 {sheet_name} 缺少列: {missing}")

        # 清洗 tags 列
        df['tags'] = df['tags'].apply(parse_tags)

        # 转换为字典列表
        data_rows = df.to_dict(orient='records')
        # print(data_rows)

        # 连接数据库并插入数据
        engine = get_database_engine(bind_key)
        Session = sessionmaker(bind=engine)
        session = Session()
        try:
            insert_data(session, model_class, data_rows)
            # print(f"插入 {data_rows}")
        except Exception as e:
            session.rollback()
            print(f"插入数据时出错: {e}")
            raise
        finally:
            session.close()

    xls.close()
    os.remove(output_file)
    print("所有数据处理完成！")

if __name__ == '__main__':
    # 备份源数据库
    timestamp = datetime.now().strftime("%Y.%m.%d_%H.%M.%S")    # 获取当前时间戳
    db_dir = Path('instance')                                   # 数据库目录
    # 重命名 navigation.db
    old_nav = db_dir / 'navigation.db'
    new_nav = db_dir / f'{timestamp}-navigation.db'
    if old_nav.exists():
        old_nav.rename(new_nav)
        print(f"已重命名: {old_nav} -> {new_nav}")
    else:
        print(f"文件不存在: {old_nav}")
    # 重命名 service.db
    old_svc = db_dir / 'service.db'
    new_svc = db_dir / f'{timestamp}-service.db'
    if old_svc.exists():
        old_svc.rename(new_svc)
        print(f"已重命名: {old_svc} -> {new_svc}")
    else:
        print(f"文件不存在: {old_svc}")

    # 写入数据
    excel_to_sqlite(r"./instance/data-human.xlsx")  # 替换为实际 Excel 文件路径